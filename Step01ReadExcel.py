from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from entity.base import Base
from entity.stock import Stock

from openpyxl import load_workbook


def init_db():
    engine = create_engine('sqlite:///./stock.db', convert_unicode=True)
    Base.metadata.create_all(engine)

    wb = load_workbook(filename="data/isino.xlsx")
    ws = wb['ISIN']

    Session = sessionmaker(bind=engine)
    session = Session()

    # From A20 onwards, A => Long Name / B => ISIN / C => Stock Code / D => Type
    row = 20
    new_stock = 0
    update_stock = 0
    while ws.cell(row=row, column=1).value is not None:
        # Check for Type
        type = ws.cell(row=row, column=4).value
        if (type != 'ORD SH') and (type != 'TRT'):
            row += 1
            continue

        stock_code = ws.cell(row=row, column=3).value
        reuter_code = "{:04}.HK".format(stock_code)
        bloomberg_code = "{:d}:HK".format(stock_code)
        isin = ws.cell(row=row, column=2).value
        fullname = ws.cell(row=row, column=1).value

        existing_stock = session.query(Stock).filter_by(reuter=reuter_code).first()

        if existing_stock is None :
            # NEW
            stock = Stock(reuter=reuter_code, bloomberg=bloomberg_code, isin=isin, fullname=fullname)
            session.add(stock)
            new_stock += 1
        else:
            # UPDATE
            existing_stock.bloomberg = bloomberg_code
            existing_stock.isin = isin
            existing_stock.fullname = fullname
            update_stock += 1

        row += 1

    print('New {0} / Update {1}'.format(new_stock, update_stock))

    session.commit()

if __name__ == '__main__':
    init_db()
